#!/usr/bin/env python3
"""
Summarize MIL ablation runs into a single Master Table.

- Recursively finds all fold_*_preds.json under ROOT.
- Computes per-sample metrics (WMAE, KS, WRMSE).
- Dynamically finds TRPOOL_*.csv logs in the `splits/` dir to EXACTLY count training 
  subset proportions without making distribution assumptions.
- Outputs an Excel file with a single sheet matching the publication table format.
"""

from __future__ import annotations

import os
import re
import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


CI_HIGH = 90

# ---------------------------------------------------------------------------
# Default Paths
# ---------------------------------------------------------------------------
IN_CONTAINER = Path("/workspace").exists()
if IN_CONTAINER:
    DEFAULT_ROOT = Path("/workspace")
else:
    DEFAULT_ROOT = Path.home() / "workspace" / "torch" / "2_1_1"

DEFAULT_PSD_XLSX = DEFAULT_ROOT / "scratch" / "labels" / "PSD.xlsx"


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("root", type=Path, help="Root directory containing ablation outputs.")
    p.add_argument(
        "--save-xlsx",
        type=Path,
        default=None,
        help="Optional path to save the summary Excel file (.xlsx). Default: <root>/summary.xlsx",
    )
    p.add_argument(
        "--psd-xlsx",
        type=Path,
        default=DEFAULT_PSD_XLSX,
        help="Path to PSD.xlsx to classify samples by size group.",
    )
    return p.parse_args(argv)


def _infer_bin_spacing(bin_x: np.ndarray) -> np.ndarray:
    if bin_x.ndim != 1:
        raise ValueError("bin_x must be one-dimensional")
    if bin_x.size == 0:
        return np.array([], dtype=float)
    if bin_x.size == 1:
        return np.array([1.0], dtype=float)
    diffs = np.diff(bin_x.astype(float))
    last = float(diffs[-1]) if diffs.size else 1.0
    return np.concatenate([diffs, [last]])


def _ensure_cdf_monotonic(cdf: np.ndarray) -> np.ndarray:
    cdf = np.clip(cdf.astype(float), 0.0, 1.0)
    return np.maximum.accumulate(cdf)


def compute_metrics(pred_cdf: np.ndarray, true_cdf: np.ndarray, bin_x: np.ndarray) -> Dict[str, float]:
    pred = _ensure_cdf_monotonic(pred_cdf)
    true = _ensure_cdf_monotonic(true_cdf)

    spacing = _infer_bin_spacing(bin_x)
    if spacing.size == 0:
        raise ValueError("Empty bin grid")

    diff = np.abs(pred - true)
    weights = spacing / float(np.sum(spacing))

    ks = float(np.max(diff))
    wmae = float(np.sum(diff * weights))
    wrmse = float(np.sqrt(np.sum((diff ** 2) * weights)))

    return {"ks": ks, "wmae": wmae, "wrmse": wrmse}


def _find_first_index(parts: Tuple[str, ...], prefix: str) -> Optional[int]:
    for i, p in enumerate(parts):
        if p.startswith(prefix):
            return i
    return None


def parse_path_keys(root: Path, preds_path: Path) -> Tuple[str, str, str, str]:
    rel = preds_path.relative_to(root)
    parts = rel.parts
    if not parts:
        raise ValueError(f"Unexpected path (empty relative): {preds_path}")

    scenario = parts[0]
    idx_seed = _find_first_index(parts, "seed_")
    idx_repeat = _find_first_index(parts, "repeat_")

    if idx_seed is not None: cut = idx_seed
    elif idx_repeat is not None: cut = idx_repeat
    else: cut = len(parts) - 1

    row_parts = parts[1:cut]
    row_id = "/".join(row_parts) if row_parts else ""

    if idx_repeat is not None: run_parts = parts[: idx_repeat + 1]
    elif idx_seed is not None: run_parts = parts[: idx_seed + 1]
    else: run_parts = parts[: len(parts) - 1]
    run_id = "/".join(run_parts)

    fold = preds_path.stem
    return scenario, row_id, run_id, fold


def iter_pred_files(root: Path) -> Iterable[Path]:
    yield from root.rglob("fold_*_preds.json")


def load_metadata(psd_path: Path) -> pd.DataFrame:
    """Loads size groups from PSD.xlsx using exactly the labels from the paper draft."""
    if not psd_path.exists():
        print(f"[WARN] PSD file not found at {psd_path}. Size groups will be 'Unknown'.")
        return pd.DataFrame(columns=["sid", "size_group"])

    df_psd = pd.read_excel(psd_path)
    if "num_publication" in df_psd.columns:
        df_psd = df_psd.rename(columns={"num_publication": "sid"})
    
    # Convert D50 to float handling commas
    if "D50" in df_psd.columns:
        df_psd["D50"] = pd.to_numeric(df_psd["D50"].astype(str).str.replace(",", "."), errors="coerce")
    else:
        df_psd["D50"] = np.nan

    def classify_by_size(row):
        d50 = row.get("D50", np.nan)
        if pd.isna(d50): return "Unknown"
        if d50 > 10:     return "Very Coarse\n(D50 > 10mm)"
        elif d50 > 2:    return "Coarse\n(D50 2-10mm)"
        elif d50 > 0.425:return "Medium\n(D50 0.4-2mm)"
        elif d50 > 0.075:return "Fine\n(D50 0.08-0.4mm)"
        else:            return "Very Fine\n(D50 < 0.08mm)"
    
    df_psd["size_group"] = df_psd.apply(classify_by_size, axis=1)
    return df_psd[["sid", "size_group"]].copy()


def collect_records(root: Path, metadata: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not root.exists():
        raise FileNotFoundError(f"Root not found: {root}")

    rows: List[Dict[str, object]] = []
    train_count_rows: List[Dict[str, object]] = []

    # Fast mapping for size lookups when we find training pools
    size_map = metadata.set_index("sid")["size_group"].to_dict() if not metadata.empty else {}
    groups_to_initialize = [
        "Very Coarse\n(D50 > 10mm)",
        "Coarse\n(D50 2-10mm)",
        "Medium\n(D50 0.4-2mm)",
        "Fine\n(D50 0.08-0.4mm)",
        "Very Fine\n(D50 < 0.08mm)",
        "Unknown"
    ]

    for preds_path in iter_pred_files(root):
        try:
            scenario, row_id, run_id, fold = parse_path_keys(root, preds_path)
        except Exception:
            continue
            
        # ---------- Exact Training Count Parsing ----------
        # 1. Read subset size `N` from `cfg.json`
        N = None
        cfg_path = preds_path.parent / "cfg.json"
        if cfg_path.exists():
            try:
                cfg_data = json.loads(cfg_path.read_text())
                N = cfg_data.get("N")
            except Exception:
                pass
                
        # Fallback to parse N from path string
        if N is None:
            n_match = re.search(r"N_(\d+)", str(preds_path))
            if n_match:
                N = int(n_match.group(1))
                
        # 2. Get the Fold ID 
        fold_idx = 0
        f_match = re.search(r"fold_(\d+)", preds_path.stem)
        if f_match:
            fold_idx = int(f_match.group(1)) - 1
            
        # 3. Find the exact `splits/` directory for this scenario
        splits_dir = None
        current = preds_path.parent
        while current != current.parent:
            if (current / "splits").is_dir():
                splits_dir = current / "splits"
                break
            current = current.parent
            
        # 4. Read the TRPOOL_X.csv to count exactly how many of each coarseness type 
        #    was included in the training subset for this fold.
        if splits_dir and N is not None:
            trpool_path = splits_dir / f"TRPOOL_{fold_idx}.csv"
            if trpool_path.exists():
                try:
                    df_tr = pd.read_csv(trpool_path)
                    # Filter to only the samples that made the N rank cut
                    train_sids = df_tr[df_tr["rank_in_trpool"] <= N]["sample_id"].tolist()
                    
                    group_counts = {grp: 0 for grp in groups_to_initialize}
                    group_counts["All samples"] = len(train_sids)
                    
                    for tsid in train_sids:
                        grp = size_map.get(tsid, "Unknown")
                        if grp in group_counts:
                            group_counts[grp] += 1
                        else:
                            group_counts["Unknown"] = group_counts.get("Unknown", 0) + 1
                            
                    for grp, cnt in group_counts.items():
                        train_count_rows.append({
                            "run_id": run_id,
                            "fold": fold,
                            "size_group": grp,
                            "train_count": cnt
                        })
                except Exception:
                    pass

        # ---------- Standard Prediction Parsing ----------
        try: payload = json.loads(preds_path.read_text())
        except json.JSONDecodeError: continue

        bin_x = np.asarray(payload.get("bin_x_log10", []), dtype=float)
        pred_list = payload.get("predictions", [])
        if bin_x.size == 0 or not pred_list:
            continue

        for rec in pred_list:
            sid = rec.get("sid", None)
            if sid is None: continue
            try: sid_int = int(sid)
            except Exception: continue

            pred_cdf = np.asarray(rec.get("pred_cdf", []), dtype=float)
            true_cdf = np.asarray(rec.get("true_cdf", []), dtype=float)
            if pred_cdf.size != true_cdf.size or pred_cdf.size != bin_x.size:
                continue

            m = compute_metrics(pred_cdf, true_cdf, bin_x)

            rows.append({
                "scenario": scenario,
                "row_id": row_id,
                "run_id": run_id,
                "sid": sid_int,
                "wmae": m["wmae"],
                "ks": m["ks"],
                "wrmse": m["wrmse"],
            })

    df = pd.DataFrame(rows)
    df_train = pd.DataFrame(train_count_rows)
    
    if df.empty:
        raise RuntimeError(f"No predictions found under: {root}")
    return df, df_train


def build_master_table(df: pd.DataFrame, metadata: pd.DataFrame, df_train_counts: pd.DataFrame) -> pd.DataFrame:
    """Builds the unified, flat table specifically requested for the paper."""
    
    # 1. Merge metadata
    if metadata.empty:
        df_merged = df.copy()
        df_merged["size_group"] = "Unknown"
    else:
        df_merged = df.merge(metadata, on="sid", how="left")
        df_merged["size_group"] = df_merged["size_group"].fillna("Unknown")

    # 2. Dynamically extract the Condition from the scenario and row_id
    def make_condition(row) -> str:
        scenario = row.get("scenario", "")
        row_id = row.get("row_id", "").replace("T224/", "")
        
        if scenario == "train_pool_sweep":
            return row_id
        if row_id:
            return f"{scenario} | {row_id}"
        return scenario

    df_merged["Condition"] = df_merged.apply(make_condition, axis=1)

    # 3. Join Condition to the exact Train Counts df so we can aggregate them
    if not df_train_counts.empty:
        run_to_cond = df_merged[["run_id", "Condition"]].drop_duplicates()
        df_train_counts = df_train_counts.merge(run_to_cond, on="run_id", how="left")

    def get_sort_tuple(cond: str) -> tuple:
        n_match = re.search(r"N_(\d+)", cond)
        n_val = int(n_match.group(1)) if n_match else 0
        
        t_match = re.search(r"valTray(\d+)", cond)
        t_val = int(t_match.group(1)) if t_match else 0
        
        base_str = re.sub(r'\d+', '', cond)
        return (base_str, t_val, n_val)

    df_merged["_sort_tup"] = df_merged["Condition"].apply(get_sort_tuple)

    # Average identical samples over folds within the same run 
    df_run_sid = df_merged.groupby(["Condition", "_sort_tup", "run_id", "sid", "size_group"], as_index=False)[["wmae", "ks", "wrmse"]].mean()

    # Define exact output group ordering
    groups_to_eval = [
        (None, "All samples"),
        ("Very Coarse\n(D50 > 10mm)", "Very Coarse\n(D50 > 10mm)"),
        ("Coarse\n(D50 2-10mm)", "Coarse\n(D50 2-10mm)"),
        ("Medium\n(D50 0.4-2mm)", "Medium\n(D50 0.4-2mm)"),
        ("Fine\n(D50 0.08-0.4mm)", "Fine\n(D50 0.08-0.4mm)"),
        ("Very Fine\n(D50 < 0.08mm)", "Very Fine\n(D50 < 0.08mm)")
    ]

    results = []
    
    for size_filter, label in groups_to_eval:
        if size_filter is None:
            group_df = df_run_sid
            tc_grp = df_train_counts[df_train_counts["size_group"] == "All samples"] if not df_train_counts.empty else pd.DataFrame()
        else:
            group_df = df_run_sid[df_run_sid["size_group"] == size_filter]
            tc_grp = df_train_counts[df_train_counts["size_group"] == size_filter] if not df_train_counts.empty else pd.DataFrame()
            
        if group_df.empty:
            continue

        for cond_val, n_df in group_df.groupby("Condition"):
            n_samples = n_df["sid"].nunique()
            sort_tup = n_df["_sort_tup"].iloc[0]
            
            # --- Get the EXACT training count average ---
            avg_train_count = np.nan
            if not tc_grp.empty:
                tc = tc_grp[tc_grp["Condition"] == cond_val]
                if not tc.empty:
                    avg_train_count = tc["train_count"].mean()
            
            if not pd.isna(avg_train_count):
                train_subset_val = int(round(avg_train_count))
            else:
                # Defensive fallback purely in case TRPOOLs were completely deleted
                n_match = re.search(r"N_(\d+)", cond_val)
                n_train_total = int(n_match.group(1)) if n_match else 0
                total_val = df_run_sid[df_run_sid["Condition"] == cond_val]["sid"].nunique()
                total_val = total_val if total_val > 0 else 1
                train_subset_val = int(round(n_train_total * (n_samples / total_val)))

            # 1. Compute median and 90th percentile across samples for EACH run independently
            run_medians = n_df.groupby("run_id")[["wmae", "ks", "wrmse"]].median()
            run_p90s = n_df.groupby("run_id")[["wmae", "ks", "wrmse"]].apply(lambda x: pd.Series({
                "wmae": np.percentile(x["wmae"], CI_HIGH),
                "ks": np.percentile(x["ks"], CI_HIGH),
                "wrmse": np.percentile(x["wrmse"], CI_HIGH)
            }))
            
            # 2. Average the run-level metrics together
            final_medians = run_medians.mean()
            final_p90s = run_p90s.mean()

            results.append({
                "Type": label,
                "Condition": cond_val,
                "Training subset\nper CV fold": train_subset_val,
                "Validation samples\n(all folds)": int(n_samples),
                "WMAE median": float(final_medians["wmae"]),
                "WMAE 90th": float(final_p90s["wmae"]),
                "KS median": float(final_medians["ks"]),
                "KS 90th": float(final_p90s["ks"]),
                "WRMSE median": float(final_medians["wrmse"]),
                "WRMSE 90th": float(final_p90s["wrmse"]),
                "_sort_tup": sort_tup 
            })

    df_res = pd.DataFrame(results)

    if not df_res.empty:
        cat_type = pd.CategoricalDtype(categories=[x[1] for x in groups_to_eval], ordered=True)
        df_res["Type"] = df_res["Type"].astype(cat_type)
        
        df_res = df_res.sort_values(["Type", "_sort_tup"]).reset_index(drop=True)
        df_res = df_res.drop(columns=["_sort_tup"])
        
        df_res["Type"] = df_res["Type"].astype(str)
        df_res.loc[df_res["Type"].duplicated(), "Type"] = ""

    return df_res


def write_excel_clean(df: pd.DataFrame, out_path: Path) -> None:
    """Writes the master table and enforces nice float display."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Master Table", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Master Table"]

    if ws.max_row > 1:
        col_map = {cell.value: cell.column for cell in ws[1]}

        for name, col_idx in col_map.items():
            if name is None: continue
            
            # Target the new column as well to ensure an integer output (0)
            if "Validation" in name or "samples" in name.lower() or "subset" in name.lower():
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "0"
            elif "WMAE" in name or "KS" in name or "WRMSE" in name:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "0.0000"

    wb.save(out_path)


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_args(argv)
    
    metadata = load_metadata(args.psd_xlsx)
    
    # Updated signature safely returns both normal predictions and precise training counts
    df, df_train = collect_records(args.root, metadata)
    master_table = build_master_table(df, metadata, df_train)

    out_path = args.save_xlsx if args.save_xlsx is not None else (args.root / "summary.xlsx")
    write_excel_clean(master_table, out_path)
    
    print(f"[SUCCESS] Saved Master Table to: {out_path}")


if __name__ == "__main__":
    main()