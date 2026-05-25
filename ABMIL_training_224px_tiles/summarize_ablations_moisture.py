#!/usr/bin/env python3
"""
Summarize MIL ablation runs into a Master Table with multiple sheets.

- Recursively finds all fold_*_preds.json under ROOT.
- Computes per-sample metrics (WMAE, KS, WRMSE).
- Outputs an Excel file with two sheets ("By Size" and "By Moisture") 
  matching the publication table format.
- Dynamically labels the ablation "Condition" to support all study plans.
"""

from __future__ import annotations

import os
import re
import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


CI_HIGH = 90

# ---------------------------------------------------------------------------
# Default Paths
# ---------------------------------------------------------------------------
IN_CONTAINER = Path("/workspace").exists()
if IN_CONTAINER:
    DEFAULT_ROOT = Path("/workspace")
    DEFAULT_PSD_XLSX = Path("/workspace/scratch/labels/PSD.xlsx")
    DEFAULT_MASS_XLSX = Path("/workspace/scratch/labels/MASS.xlsx")
else:
    DEFAULT_ROOT = Path("/home/thomas_plante_stcyr/workspace/torch/2_1_1")
    DEFAULT_PSD_XLSX = DEFAULT_ROOT / "scratch" / "labels" / "PSD.xlsx"
    DEFAULT_MASS_XLSX = DEFAULT_ROOT / "scratch" / "labels" / "MASS.xlsx"


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("root", type=Path, nargs="?", default=DEFAULT_ROOT, help="Root directory containing ablation outputs.")
    p.add_argument(
        "--save-xlsx",
        type=Path,
        default=None,
        help="Optional path to save the summary Excel file (.xlsx). Default: <root>/summary.xlsx",
    )
    p.add_argument(
        "--psd-xlsx",
        type=Path,
        default=DEFAULT_PSD_XLSX,
        help="Path to PSD.xlsx to classify samples by size.",
    )
    p.add_argument(
        "--mass-xlsx",
        type=Path,
        default=DEFAULT_MASS_XLSX,
        help="Path to MASS.xlsx to classify samples by moisture.",
    )
    return p.parse_args(argv)


def _infer_bin_spacing(bin_x: np.ndarray) -> np.ndarray:
    if bin_x.ndim != 1:
        raise ValueError("bin_x must be one-dimensional")
    if bin_x.size == 0:
        return np.array([], dtype=float)
    if bin_x.size == 1:
        return np.array([1.0], dtype=float)
    diffs = np.diff(bin_x.astype(float))
    last = float(diffs[-1]) if diffs.size else 1.0
    return np.concatenate([diffs, [last]])


def _ensure_cdf_monotonic(cdf: np.ndarray) -> np.ndarray:
    cdf = np.clip(cdf.astype(float), 0.0, 1.0)
    return np.maximum.accumulate(cdf)


def compute_metrics(pred_cdf: np.ndarray, true_cdf: np.ndarray, bin_x: np.ndarray) -> Dict[str, float]:
    pred = _ensure_cdf_monotonic(pred_cdf)
    true = _ensure_cdf_monotonic(true_cdf)

    spacing = _infer_bin_spacing(bin_x)
    if spacing.size == 0:
        raise ValueError("Empty bin grid")

    diff = np.abs(pred - true)
    weights = spacing / float(np.sum(spacing))

    ks = float(np.max(diff))
    wmae = float(np.sum(diff * weights))
    wrmse = float(np.sqrt(np.sum((diff ** 2) * weights)))

    return {"ks": ks, "wmae": wmae, "wrmse": wrmse}


def _find_first_index(parts: Tuple[str, ...], prefix: str) -> Optional[int]:
    for i, p in enumerate(parts):
        if p.startswith(prefix):
            return i
    return None


def parse_path_keys(root: Path, preds_path: Path) -> Tuple[str, str, str, str]:
    rel = preds_path.relative_to(root)
    parts = rel.parts
    if not parts:
        raise ValueError(f"Unexpected path (empty relative): {preds_path}")

    scenario = parts[0]
    idx_seed = _find_first_index(parts, "seed_")
    idx_repeat = _find_first_index(parts, "repeat_")

    if idx_seed is not None: cut = idx_seed
    elif idx_repeat is not None: cut = idx_repeat
    else: cut = len(parts) - 1

    row_parts = parts[1:cut]
    row_id = "/".join(row_parts) if row_parts else ""

    if idx_repeat is not None: run_parts = parts[: idx_repeat + 1]
    elif idx_seed is not None: run_parts = parts[: idx_seed + 1]
    else: run_parts = parts[: len(parts) - 1]
    run_id = "/".join(run_parts)

    fold = preds_path.stem
    return scenario, row_id, run_id, fold


def iter_pred_files(root: Path) -> Iterable[Path]:
    yield from root.rglob("fold_*_preds.json")


def load_metadata(psd_path: Path, mass_path: Path) -> pd.DataFrame:
    """Loads size groups from PSD.xlsx and moisture groups from MASS.xlsx."""
    
    # --- 1. Load Size Data from PSD.xlsx ---
    if not psd_path.exists():
        print(f"[WARN] PSD file not found at {psd_path}. Size groups will be 'Unknown'.")
        df_size = pd.DataFrame(columns=["sid", "size_group"])
    else:
        df_psd = pd.read_excel(psd_path)
        if "num_publication" in df_psd.columns:
            df_psd = df_psd.rename(columns={"num_publication": "sid"})
        
        if "D50" in df_psd.columns:
            df_psd["D50"] = pd.to_numeric(df_psd["D50"].astype(str).str.replace(",", "."), errors="coerce")
        else:
            df_psd["D50"] = np.nan

        def classify_by_size(row):
            d50 = row.get("D50", np.nan)
            if pd.isna(d50): return "Unknown"
            if d50 > 10:     return "Very Coarse\n(D50 > 10mm)"
            elif d50 > 2:    return "Coarse\n(D50 2-10mm)"
            elif d50 > 0.425:return "Medium\n(D50 0.4-2mm)"
            elif d50 > 0.075:return "Fine\n(D50 0.08-0.4mm)"
            else:            return "Very Fine\n(D50 < 0.08mm)"
        
        df_psd["size_group"] = df_psd.apply(classify_by_size, axis=1)
        df_size = df_psd[["sid", "size_group"]].copy()

    # --- 2. Load Moisture Data from MASS.xlsx ---
    if not mass_path.exists():
        print(f"[WARN] MASS file not found at {mass_path}. Moisture groups will be 'Unknown'.")
        df_moisture = pd.DataFrame(columns=["sid", "moisture_group"])
    else:
        df_mass = pd.read_excel(mass_path)
        if "num_publication" in df_mass.columns:
            df_mass = df_mass.rename(columns={"num_publication": "sid"})

        def classify_by_moisture(row):
            water_pct = row.get("%_water", np.nan)
            if pd.isna(water_pct): return "Unknown"
            if water_pct < 5.0:    return "< 5% Moisture"
            elif water_pct < 10.0: return "5 - 10% Moisture"
            elif water_pct < 15.0: return "10 - 15% Moisture"
            elif water_pct < 20.0: return "15 - 20% Moisture"
            else:                  return "> 20% Moisture"
        
        if "%_water" in df_mass.columns:
            df_mass["moisture_group"] = df_mass.apply(classify_by_moisture, axis=1)
        else:
            print("[WARN] Column '%_water' not found in MASS.xlsx. All will be 'Unknown'.")
            df_mass["moisture_group"] = "Unknown"
            
        df_moisture = df_mass[["sid", "moisture_group"]].copy()

    # --- Merge the two metadata tables together ---
    if df_size.empty and df_moisture.empty:
        return pd.DataFrame(columns=["sid", "size_group", "moisture_group"])
    
    if df_size.empty:
        df_size = pd.DataFrame({"sid": df_moisture["sid"], "size_group": "Unknown"})
    if df_moisture.empty:
        df_moisture = pd.DataFrame({"sid": df_size["sid"], "moisture_group": "Unknown"})

    metadata = pd.merge(df_size, df_moisture, on="sid", how="outer").fillna("Unknown")
    return metadata


def collect_records(root: Path) -> pd.DataFrame:
    if not root.exists():
        raise FileNotFoundError(f"Root not found: {root}")

    rows: List[Dict[str, object]] = []

    for preds_path in iter_pred_files(root):
        try:
            scenario, row_id, run_id, fold = parse_path_keys(root, preds_path)
        except Exception:
            continue

        try: payload = json.loads(preds_path.read_text())
        except json.JSONDecodeError: continue

        bin_x = np.asarray(payload.get("bin_x_log10", []), dtype=float)
        pred_list = payload.get("predictions", [])
        if bin_x.size == 0 or not pred_list:
            continue

        for rec in pred_list:
            sid = rec.get("sid", None)
            if sid is None: continue
            try: sid_int = int(sid)
            except Exception: continue

            pred_cdf = np.asarray(rec.get("pred_cdf", []), dtype=float)
            true_cdf = np.asarray(rec.get("true_cdf", []), dtype=float)
            if pred_cdf.size != true_cdf.size or pred_cdf.size != bin_x.size:
                continue

            m = compute_metrics(pred_cdf, true_cdf, bin_x)

            rows.append({
                "scenario": scenario,
                "row_id": row_id,
                "run_id": run_id,
                "sid": sid_int,
                "wmae": m["wmae"],
                "ks": m["ks"],
                "wrmse": m["wrmse"],
            })

    df = pd.DataFrame(rows)
    if df.empty:
        raise RuntimeError(f"No predictions found under: {root}")
    return df


def build_summary_table(df: pd.DataFrame, metadata: pd.DataFrame, group_col: str, groups_to_eval: List[Tuple[Optional[str], str]]) -> pd.DataFrame:
    """Builds a unified, flat table grouped by the specified column."""
    
    # 1. Merge metadata
    if metadata.empty:
        df_merged = df.copy()
        df_merged[group_col] = "Unknown"
    else:
        df_merged = df.merge(metadata, on="sid", how="left")
        df_merged[group_col] = df_merged[group_col].fillna("Unknown")

    # 2. Dynamically extract the Condition
    def make_condition(row) -> str:
        scenario = row.get("scenario", "")
        row_id = row.get("row_id", "").replace("T224/", "")
        
        if scenario == "train_pool_sweep":
            return row_id
        if row_id:
            return f"{scenario} | {row_id}"
        return scenario

    df_merged["Condition"] = df_merged.apply(make_condition, axis=1)

    # Helper function for sorting conditions
    def get_sort_tuple(cond: str) -> tuple:
        n_match = re.search(r"N_(\d+)", cond)
        n_val = int(n_match.group(1)) if n_match else 0
        
        t_match = re.search(r"valTray(\d+)", cond)
        t_val = int(t_match.group(1)) if t_match else 0
        
        base_str = re.sub(r'\d+', '', cond)
        return (base_str, t_val, n_val)

    df_merged["_sort_tup"] = df_merged["Condition"].apply(get_sort_tuple)

    # Average identical samples over folds within the same run 
    df_run_sid = df_merged.groupby(["Condition", "_sort_tup", "run_id", "sid", group_col], as_index=False)[["wmae", "ks", "wrmse"]].mean()

    results = []
    
    for filter_val, label in groups_to_eval:
        if filter_val is None:
            group_df = df_run_sid
        else:
            group_df = df_run_sid[df_run_sid[group_col] == filter_val]
            
        if group_df.empty:
            continue

        for cond_val, n_df in group_df.groupby("Condition"):
            n_samples = n_df["sid"].nunique()
            sort_tup = n_df["_sort_tup"].iloc[0]
            
            run_medians = n_df.groupby("run_id")[["wmae", "ks", "wrmse"]].median()
            run_p90s = n_df.groupby("run_id")[["wmae", "ks", "wrmse"]].apply(lambda x: pd.Series({
                "wmae": np.percentile(x["wmae"], CI_HIGH),
                "ks": np.percentile(x["ks"], CI_HIGH),
                "wrmse": np.percentile(x["wrmse"], CI_HIGH)
            }))
            
            final_medians = run_medians.mean()
            final_p90s = run_p90s.mean()

            results.append({
                "Type": label,
                "Condition": cond_val,
                "Validation samples\n(all folds)": int(n_samples),
                "WMAE median": float(final_medians["wmae"]),
                "WMAE 90th": float(final_p90s["wmae"]),
                "KS median": float(final_medians["ks"]),
                "KS 90th": float(final_p90s["ks"]),
                "WRMSE median": float(final_medians["wrmse"]),
                "WRMSE 90th": float(final_p90s["wrmse"]),
                "_sort_tup": sort_tup
            })

    df_res = pd.DataFrame(results)

    if not df_res.empty:
        # Enforce ordering based on the passed groups
        cat_type = pd.CategoricalDtype(categories=[x[1] for x in groups_to_eval], ordered=True)
        df_res["Type"] = df_res["Type"].astype(cat_type)
        
        df_res = df_res.sort_values(["Type", "_sort_tup"]).reset_index(drop=True)
        df_res = df_res.drop(columns=["_sort_tup"])
        
        # Visual trick: Blank out duplicated "Type" names
        df_res["Type"] = df_res["Type"].astype(str)
        df_res.loc[df_res["Type"].duplicated(), "Type"] = ""

    return df_res


def write_excel_clean(sheets_dict: Dict[str, pd.DataFrame], out_path: Path) -> None:
    """Writes multiple tables to separate sheets and enforces float display."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)

    for sheet_name in sheets_dict.keys():
        ws = wb[sheet_name]

        if ws.max_row > 1:
            col_map = {cell.value: cell.column for cell in ws[1]}

            for name, col_idx in col_map.items():
                if name is None: continue
                
                if "Validation" in name or "samples" in name.lower():
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).number_format = "0"
                elif "WMAE" in name or "KS" in name or "WRMSE" in name:
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).number_format = "0.0000"

    wb.save(out_path)


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_args(argv)
    
    # Load and process data
    df = collect_records(args.root)
    metadata = load_metadata(args.psd_xlsx, args.mass_xlsx)
    
    # --- 1. Build By Size Table ---
    groups_size = [
        (None, "All samples"),
        ("Very Coarse\n(D50 > 10mm)", "Very Coarse\n(D50 > 10mm)"),
        ("Coarse\n(D50 2-10mm)", "Coarse\n(D50 2-10mm)"),
        ("Medium\n(D50 0.4-2mm)", "Medium\n(D50 0.4-2mm)"),
        ("Fine\n(D50 0.08-0.4mm)", "Fine\n(D50 0.08-0.4mm)"),
        ("Very Fine\n(D50 < 0.08mm)", "Very Fine\n(D50 < 0.08mm)")
    ]
    table_size = build_summary_table(df, metadata, "size_group", groups_size)

    # --- 2. Build By Moisture Table ---
    groups_moisture = [
        (None, "All samples"),
        ("< 5% Moisture", "< 5% Moisture"),
        ("5 - 10% Moisture", "5 - 10% Moisture"),
        ("10 - 15% Moisture", "10 - 15% Moisture"),
        ("15 - 20% Moisture", "15 - 20% Moisture"),
        ("> 20% Moisture", "> 20% Moisture"),
    ]
    table_moisture = build_summary_table(df, metadata, "moisture_group", groups_moisture)

    # Save
    out_path = args.save_xlsx if args.save_xlsx is not None else (args.root / "summary_moisture.xlsx")
    sheets = {
        "By Size": table_size,
        "By Moisture": table_moisture
    }
    write_excel_clean(sheets, out_path)
    
    print(f"[SUCCESS] Saved Master Tables to: {out_path}")


if __name__ == "__main__":
    main()